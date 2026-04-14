import uuid
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta, timezone
from fastapi import APIRouter, HTTPException, Depends, Query, BackgroundTasks, Response
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.future import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import func, cast, Date, update
from pydantic import BaseModel

from apps.api.src.services.auth.deps import admin_guard
from apps.api.src.db.session import get_db
from apps.api.src.schemas.payments import (
    PackageCreate, PackageUpdate, CouponCreate, CouponUpdate,
    PackageResponse, CouponResponse, AdminTransactionResponse
)
from sqlalchemy.orm import joinedload
from apps.api.src.schemas.user import UserResponseAdmin, UserCreateAdmin
from apps.api.src.services.auth.utils import get_password_hash
from apps.api.src.db.models.base import UserProfile, UserUsage, User, ChatSession, ChatMessage, CreditPackage, PaymentTransaction, Coupon, CreditLog
from apps.api.src.services.payments import initialize_user_credits, assign_invoice_number, send_invoice_background

router = APIRouter(prefix="/admin", tags=["Admin"])

# --- User Management ---
class UserUpdateAdmin(BaseModel):
    role: Optional[str] = None
    max_sessions: Optional[int] = None
    password: Optional[str] = None
    is_verified: Optional[bool] = None
    full_name: Optional[str] = None
    mobile_number: Optional[str] = None
    state: Optional[str] = None
    gst_number: Optional[str] = None

class UserUsageUpdateAdmin(BaseModel):
    simple_query_balance: Optional[int] = None
    draft_reply_balance: Optional[int] = None
    reset_monthly_tokens: Optional[bool] = False

@router.get("/users", response_model=List[UserResponseAdmin])
async def list_users(admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User))
    return result.scalars().all()

@router.post("/users", response_model=UserResponseAdmin)
async def create_user_admin(payload: UserCreateAdmin, background_tasks: BackgroundTasks, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    email = payload.email.lower()
    # Check if exists
    result = await db.execute(select(User).where(func.lower(User.email) == email))
    if result.scalars().first():
        raise HTTPException(status_code=400, detail="Email already registered")
    
    new_user = User(
        email=email,
        password_hash=get_password_hash(payload.password),
        full_name=payload.full_name,
        mobile_number=payload.mobile_number,
        state=payload.state,
        gst_number=payload.gst_number,
        country=payload.country,
        role=payload.role,
        is_verified=payload.is_verified,
        referral_code=payload.referral_code,
        onboarding_step=2 if payload.package_id else 1
    )
    if payload.max_sessions is not None:
        new_user.max_sessions = payload.max_sessions
        
    db.add(new_user)
    await db.commit()
    await db.refresh(new_user)

    # Initialize profile
    db.add(UserProfile(user_id=new_user.id))
    
    # Initialize credits based on selection
    order_id = None
    if payload.package_id and payload.base_amount is not None:
        # Fetch individual package details
        pkg_res = await db.execute(select(CreditPackage).where(CreditPackage.id == payload.package_id))
        pkg = pkg_res.scalars().first()
        if not pkg:
            raise HTTPException(status_code=400, detail="Invalid package ID")

        # calculate final amount with GST (18%)
        base_amount = payload.base_amount
        gst_amount = round(base_amount * 0.18, 2)
        total_amount = round(base_amount + gst_amount, 2)

        # Create Payment Transaction (Admin Assigned)
        order_id = f"admin_{uuid.uuid4().hex[:8]}"
        transaction = PaymentTransaction(
            user_id=new_user.id, order_id=order_id, payment_id="admin_assigned",
            amount=int(total_amount), # Store in paise/cents
            package_id=pkg.id, 
            draft_credits_added=pkg.draft_credits,
            simple_credits_added=pkg.simple_credits,
            user_gst_number=payload.gst_number,
            status="completed"
        )
        db.add(transaction)
        await db.flush() # Get ID for invoice
        await assign_invoice_number(db, transaction)

        # Update Usage
        usage = UserUsage(
            user_id=new_user.id,
            simple_query_balance=pkg.simple_credits,
            draft_reply_balance=pkg.draft_credits,
            credits_expire_at=datetime.now(timezone.utc) + timedelta(days=pkg.validity_days or 365)
        )
        db.add(usage)

        # Log credits
        db.add(CreditLog(
            user_id=new_user.id, amount=pkg.simple_credits, credit_type="simple",
            transaction_type="purchase", reference_id=order_id
        ))
        db.add(CreditLog(
            user_id=new_user.id, amount=pkg.draft_credits, credit_type="draft",
            transaction_type="purchase", reference_id=order_id
        ))
    else:
        # Initialize with zero credits (Step 1)
        await initialize_user_credits(new_user.id, db, use_welcome_package=False)
    
    await db.commit()
    await db.refresh(new_user)
    
    # Send invoice email if a package was assigned
    if payload.package_id and payload.base_amount is not None:
        background_tasks.add_task(send_invoice_background, order_id=order_id)

    return new_user

@router.get("/users/{user_id}", response_model=UserResponseAdmin)
async def get_user(user_id: int, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalars().first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@router.patch("/users/{user_id}", response_model=UserResponseAdmin)
async def update_user(user_id: int, payload: UserUpdateAdmin, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalars().first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if payload.role is not None:
        user.role = payload.role
    if payload.max_sessions is not None:
        user.max_sessions = payload.max_sessions
    if payload.password is not None:
        user.password_hash = get_password_hash(payload.password)
    if payload.is_verified is not None:
        user.is_verified = payload.is_verified
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.mobile_number is not None:
        user.mobile_number = payload.mobile_number
    if payload.state is not None:
        user.state = payload.state
    if payload.gst_number is not None:
        user.gst_number = payload.gst_number
    
    await db.commit()
    await db.refresh(user)
    return user

@router.get("/users/{user_id}/usage")
async def get_user_usage(
    user_id: int, 
    admin_user=Depends(admin_guard), 
    db: AsyncSession = Depends(get_db)
):
    """
    Get the usage and token balances for a specific user.
    """
    result = await db.execute(select(UserUsage).where(UserUsage.user_id == user_id))
    usage = result.scalars().first()
    
    if not usage:
        return {
            "user_id": user_id,
            "simple_query_balance": 1000000, 
            "draft_reply_balance": 3,
            "monthly_tokens_used": 0,
            "monthly_reset_date": datetime.now(timezone.utc)
        }
    
    return {
        "user_id": usage.user_id,
        "simple_query_balance": usage.simple_query_balance,
        "draft_reply_balance": usage.draft_reply_balance,
        "monthly_tokens_used": usage.monthly_tokens_used,
        "monthly_reset_date": usage.monthly_reset_date
    }

@router.patch("/users/{user_id}/usage")
async def update_user_usage(
    user_id: int, 
    payload: UserUsageUpdateAdmin, 
    admin_user=Depends(admin_guard), 
    db: AsyncSession = Depends(get_db)
):
    """
    Allow admins to manually override balances or force a monthly FUP token reset.
    """
    result = await db.execute(select(UserUsage).where(UserUsage.user_id == user_id))
    usage = result.scalars().first()
    
    if not usage:
        usage = UserUsage(user_id=user_id)
        db.add(usage)
        
    if payload.simple_query_balance is not None:
        usage.simple_query_balance = payload.simple_query_balance
    if payload.draft_reply_balance is not None:
        usage.draft_reply_balance = payload.draft_reply_balance
    if payload.reset_monthly_tokens:
        usage.monthly_tokens_used = 0
        usage.monthly_reset_date = datetime.now(timezone.utc)
        
    await db.commit()
    await db.refresh(usage)
    
    return {
        "user_id": usage.user_id,
        "simple_query_balance": usage.simple_query_balance,
        "draft_reply_balance": usage.draft_reply_balance,
        "monthly_tokens_used": usage.monthly_tokens_used,
        "monthly_reset_date": usage.monthly_reset_date
    }

@router.delete("/users/{user_id}")
async def delete_user(user_id: int, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalars().first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    await db.delete(user)
    await db.commit()
    return {"status": "success"}

# --- Package & Coupon Management ---
@router.post("/payments/packages", response_model=PackageResponse)
async def create_package(payload: PackageCreate, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    payload_dict = payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()
    
    # If this package is being set as default, unset others
    if payload_dict.get("is_default"):
        await db.execute(update(CreditPackage).values(is_default=False).where(CreditPackage.is_default == True))
        
    new_pkg = CreditPackage(**payload_dict)
    db.add(new_pkg)
    await db.commit()
    await db.refresh(new_pkg)
    return new_pkg

@router.get("/payments/packages", response_model=List[PackageResponse])
async def list_all_packages(admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(CreditPackage).where(CreditPackage.is_deleted == False))
    return res.scalars().all()

@router.post("/payments/coupons", response_model=CouponResponse)
async def create_coupon(payload: CouponCreate, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    payload_dict = payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()
    new_coupon = Coupon(**payload_dict)
    db.add(new_coupon)
    await db.commit()
    await db.refresh(new_coupon)
    return new_coupon

@router.get("/payments/transactions", response_model=List[AdminTransactionResponse])
async def list_all_transactions(admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    """Get all payment transactions with package and user details across the platform."""
    res = await db.execute(
        select(PaymentTransaction)
        .options(joinedload(PaymentTransaction.package), joinedload(PaymentTransaction.user))
        .order_by(PaymentTransaction.created_at.desc())
        .limit(200)
    )
    return res.scalars().all()

@router.get("/analytics")
async def get_analytics(
    user_id: Optional[int] = None,
    admin_user=Depends(admin_guard),
    db: AsyncSession = Depends(get_db)
):
    """
    Overall platform stats. If user_id is provided, scopes today/yesterday
    usage to that specific user and omits total_users from the response.
    """
    now = datetime.now(timezone.utc)
    today = now.date()
    yesterday = today - timedelta(days=1)

    # Revenue & sales from completed transactions
    revenue_q = select(func.coalesce(func.sum(PaymentTransaction.amount), 0)).where(
        PaymentTransaction.status == "completed"
    )
    sales_q = select(func.count()).select_from(PaymentTransaction).where(
        PaymentTransaction.status == "completed"
    )
    if user_id is not None:
        revenue_q = revenue_q.where(PaymentTransaction.user_id == user_id)
        sales_q = sales_q.where(PaymentTransaction.user_id == user_id)

    total_revenue_paise = await db.scalar(revenue_q)
    total_sales = await db.scalar(sales_q)

    # Today / yesterday query counts from ChatMessage (role='user' = one query)
    today_q = (
        select(func.count())
        .select_from(ChatMessage)
        .join(ChatSession, ChatMessage.session_id == ChatSession.id)
        .where(
            ChatMessage.role == "user",
            cast(ChatMessage.timestamp, Date) == today
        )
    )
    yesterday_q = (
        select(func.count())
        .select_from(ChatMessage)
        .join(ChatSession, ChatMessage.session_id == ChatSession.id)
        .where(
            ChatMessage.role == "user",
            cast(ChatMessage.timestamp, Date) == yesterday
        )
    )
    if user_id is not None:
        today_q = today_q.where(ChatSession.user_id == user_id)
        yesterday_q = yesterday_q.where(ChatSession.user_id == user_id)

    today_usage = await db.scalar(today_q)
    yesterday_usage = await db.scalar(yesterday_q)

    response: Dict[str, Any] = {
        "total_revenue_paise": total_revenue_paise or 0,
        "total_sales": total_sales or 0,
        "today_usage": today_usage or 0,
        "yesterday_usage": yesterday_usage or 0,
    }

    # Only include total_users when not scoped to a specific user
    if user_id is None:
        response["total_users"] = await db.scalar(select(func.count()).select_from(User))

    return response


@router.get("/analytics/users")
async def top_users_today(
    limit: int = 20,
    admin_user=Depends(admin_guard),
    db: AsyncSession = Depends(get_db)
):
    """
    Returns users ranked by number of queries made today, with today and
    yesterday usage counts included.
    """
    now = datetime.now(timezone.utc)
    today = now.date()
    yesterday = today - timedelta(days=1)

    # Subquery: today's message counts per user
    today_sub = (
        select(
            ChatSession.user_id,
            func.count(ChatMessage.id).label("today_usage")
        )
        .join(ChatMessage, ChatSession.id == ChatMessage.session_id)
        .where(
            ChatMessage.role == "user",
            cast(ChatMessage.timestamp, Date) == today
        )
        .group_by(ChatSession.user_id)
        .subquery()
    )

    # Subquery: yesterday's message counts per user
    yesterday_sub = (
        select(
            ChatSession.user_id,
            func.count(ChatMessage.id).label("yesterday_usage")
        )
        .join(ChatMessage, ChatSession.id == ChatMessage.session_id)
        .where(
            ChatMessage.role == "user",
            cast(ChatMessage.timestamp, Date) == yesterday
        )
        .group_by(ChatSession.user_id)
        .subquery()
    )

    res = await db.execute(
        select(
            User.id,
            User.full_name,
            User.email,
            func.coalesce(today_sub.c.today_usage, 0).label("today_usage"),
            func.coalesce(yesterday_sub.c.yesterday_usage, 0).label("yesterday_usage"),
        )
        .outerjoin(today_sub, User.id == today_sub.c.user_id)
        .outerjoin(yesterday_sub, User.id == yesterday_sub.c.user_id)
        .order_by(func.coalesce(today_sub.c.today_usage, 0).desc())
        .limit(limit)
    )

    return [
        {
            "user_id": row.id,
            "full_name": row.full_name,
            "email": row.email,
            "today_usage": row.today_usage,
            "yesterday_usage": row.yesterday_usage,
        }
        for row in res.all()
    ]

@router.patch("/payments/packages/{package_id}", response_model=PackageResponse)
async def update_package(package_id: int, payload: PackageUpdate, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(CreditPackage).where(CreditPackage.id == package_id))
    pkg = res.scalars().first()
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found")
    
    update_data = payload.model_dump(exclude_unset=True) if hasattr(payload, "model_dump") else payload.dict(exclude_unset=True)
    
    # If this package is being set as default, unset others
    if update_data.get("is_default"):
        await db.execute(update(CreditPackage).values(is_default=False).where(CreditPackage.id != package_id, CreditPackage.is_default == True))
        
    for k, v in update_data.items():
        setattr(pkg, k, v)
    
    await db.commit()
    await db.refresh(pkg)
    return pkg

@router.delete("/payments/packages/{package_id}")
async def delete_package(package_id: int, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(CreditPackage).where(CreditPackage.id == package_id))
    pkg = res.scalars().first()
    if not pkg:
        raise HTTPException(status_code=404, detail="Package not found")
    
    # Check if package is used in any transactions
    txn_res = await db.execute(select(PaymentTransaction).where(PaymentTransaction.package_id == package_id).limit(1))
    has_transactions = txn_res.scalars().first() is not None

    if has_transactions:
        # Soft delete
        pkg.is_deleted = True
        pkg.is_active = False
        pkg.name = f"{pkg.name}_deleted_{int(datetime.now(timezone.utc).timestamp())}"
        await db.commit()
        return {"status": "archived", "message": "Package archived as it has transaction history"}
    
    # Hard delete
    await db.delete(pkg)
    await db.commit()
    return {"status": "deleted"}

@router.get("/payments/coupons", response_model=List[CouponResponse])
async def list_coupons(admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Coupon).where(Coupon.is_deleted == False))
    return res.scalars().all()

@router.patch("/payments/coupons/{coupon_id}", response_model=CouponResponse)
async def update_coupon(coupon_id: int, payload: CouponUpdate, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Coupon).where(Coupon.id == coupon_id))
    coupon = res.scalars().first()
    if not coupon:
        raise HTTPException(status_code=404, detail="Coupon not found")
    
    update_data = payload.model_dump(exclude_unset=True) if hasattr(payload, "model_dump") else payload.dict(exclude_unset=True)
    for k, v in update_data.items():
        setattr(coupon, k, v)
    
    await db.commit()
    await db.refresh(coupon)
    return coupon

@router.delete("/payments/coupons/{coupon_id}")
async def delete_coupon(coupon_id: int, admin_user=Depends(admin_guard), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Coupon).where(Coupon.id == coupon_id))
    coupon = res.scalars().first()
    if not coupon:
        raise HTTPException(status_code=404, detail="Coupon not found")
    
    # Check if coupon is used in any transactions
    txn_res = await db.execute(select(PaymentTransaction).where(PaymentTransaction.coupon_id == coupon_id).limit(1))
    has_transactions = txn_res.scalars().first() is not None

    if has_transactions:
        # Soft delete
        coupon.is_deleted = True
        coupon.is_active = False
        coupon.code = f"{coupon.code}_deleted_{int(datetime.now(timezone.utc).timestamp())}"
        await db.commit()
        return {"status": "archived", "message": "Coupon archived as it has transaction history"}
    
    # Hard delete
    await db.delete(coupon)
    await db.commit()
    return {"status": "deleted"}

@router.get("/reports/mis/download")
async def download_mis_report(
    admin_user=Depends(admin_guard),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate and download an Excel MIS report of all users and their latest transaction details.
    """
    # Fetch all users with their usage and transactions
    # We order transactions by created_at desc to easily pick the latest one
    result = await db.execute(
        select(User)
        .options(
            joinedload(User.usage),
            joinedload(User.transactions).joinedload(PaymentTransaction.package)
        )
    )
    users = result.unique().scalars().all()

    # Create a new Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User MIS Report"

    # Define Columns
    headers = [
        "User Fullname", "Email", "Phone Number", "Referral Code", 
        "User Created At", "Last Login At", "Package Expiry", 
        "Last Transaction Date", "Last Order ID", "Last Amount (INR)", 
        "Last Package Name", "Last GSTIN", "Queries Used", "Drafts Used"
    ]

    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Populate Data
    for row_num, user in enumerate(users, 2):
        # Find the latest completed transaction
        completed_txns = sorted(
            [t for t in user.transactions if t.status == "completed"],
            key=lambda x: x.created_at,
            reverse=True
        )
        last_txn = completed_txns[0] if completed_txns else None
        
        usage = user.usage
        
        data = [
            user.full_name or "N/A",
            user.email,
            user.mobile_number or "N/A",
            user.referral_code or "N/A",
            user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "N/A",
            user.last_login_at.strftime("%Y-%m-%d %H:%M") if user.last_login_at else "N/A",
            usage.credits_expire_at.strftime("%Y-%m-%d") if usage and usage.credits_expire_at else "N/A",
            last_txn.created_at.strftime("%Y-%m-%d %H:%M") if last_txn else "N/A",
            last_txn.order_id if last_txn else "N/A",
            (last_txn.amount / 100) if last_txn else 0, # Convert paise to INR
            last_txn.package.title if last_txn and last_txn.package else "N/A",
            last_txn.user_gst_number if last_txn else (user.gst_number or "N/A"),
            usage.simple_query_used if usage else 0,
            usage.draft_reply_used if usage else 0
        ]
        
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num).value = value

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"MIS_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
